import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

import ast

def get_category():
    res_df = pd.read_excel('export_sakura2.xlsx')
    # Убираем дубликаты по Марке, Модели двигателю и Номеру артикля
    res_df = res_df.drop_duplicates(subset=['Name', 'VM', "ArticleNumber", "Engines", "TypeName"])
    #
    name_df = res_df[["Name", "VM", "Engines", "TypeName", "HorsePowers", "Year"]].drop_duplicates(subset=['Name', 'VM', "Engines", "TypeName"])
    # print(name_df)

    # Создаем список всех категорий артиклей
    articles = res_df[["GenericArticle"]].drop_duplicates(subset=['GenericArticle'])['GenericArticle'].tolist()

    # Из списка артиклей создаем поля
    for article in articles:
        res_df.insert(loc=len(name_df.columns), column=article, value=article)

    for i, row in res_df.iterrows():
        article = row["GenericArticle"]
        res_df.loc[i, row[article]] = row['ArticleNumber']
    res_df = res_df.drop_duplicates(subset=['Name', 'VM', "Engines", "Year", "HorsePowers", 'ArticleNumber'])

    # Проходим по начальному датафрейму и заполняем созданные столбцы артиклей
    data_dicts = res_df.to_dict('records')
    for d in data_dicts:
        for x in d.keys():
            if d[x] == x:
                d[x] = '*'
    result_df = pd.DataFrame(data_dicts)

    # Группируем артикли по 'Name', "VM", "Engines"
    result_df = result_df.groupby(['Name', "VM", "Engines"])[articles].agg(','.join).reset_index()
    r_df = name_df.merge(result_df, how='inner', on=["Name", "VM", "Engines"])
    r_df = r_df.reset_index()

    # Очищаем ячейки от * и ','
    for art in articles:
        r_df[art] = r_df[art].str.replace('*,', '')
        r_df[art] = r_df[art].str.replace(',*', '')
        r_df[art] = r_df[art].str.replace('*', '')

    # Проверка на дубликаты в ячейках
    for i, row in r_df.iterrows():
        for art in articles:
            if row[art]:
                x = len(list(set(row[art].split(','))))
                y = len(row[art].split(','))
                if x != y:
                    print(x, y)

    r_df = r_df.sort_values(['Name', 'VM'])
    r_df = r_df.drop(columns=['index'])
    print(r_df)

    r_df.to_excel("result_colum_category2.xlsx", index=False)


def change_colum():

    df = pd.read_excel('result_colum_category.xlsx')
    brand_dict = {}

    for i, row in df.iterrows():
        brand = row['Name']
        model = row['VM']
        engine = row['Engines']
        engine_capacity = row['TypeName']
        hp = row['HorsePowers']
        year = row['Year']

        if brand not in brand_dict:
            brand_dict[brand] = {}
        if model not in brand_dict[brand]:
            brand_dict[brand][model] = {}
            if '-' in year:
                year_start, year_end = year.split('-')
                if year_start != '':
                    brand_dict[brand][model]["start_date"] = year_start
                else:
                    brand_dict[brand][model]["start_date"] = year
                if year_end != '':
                    brand_dict[brand][model]["end_date"] = year_end
                else:
                    brand_dict[brand][model]["end_date"] = year
            else:
                year_start = year_end = year
                brand_dict[brand][model]["start_date"] = year_start
                brand_dict[brand][model]["end_date"] = year_end
        else:
            if '-' in year:
                year_start, year_end = year.split('-')
                if year_start != '' and year_start < brand_dict[brand][model]["start_date"]:
                    brand_dict[brand][model]["start_date"] = year_start
                if year_end != '' and year_end > brand_dict[brand][model]["end_date"]:
                    brand_dict[brand][model]["end_date"] = year_end
            else:
                year_start = year_end = year
                if year_start < brand_dict[brand][model]["start_date"]:
                    brand_dict[brand][model]["start_date"] = year_start
                if year_end > brand_dict[brand][model]["end_date"]:
                    brand_dict[brand][model]["end_date"] = year_end

        if engine_capacity not in brand_dict[brand][model]:
            brand_dict[brand][model][engine_capacity] = {}
        if engine not in brand_dict[brand][model][engine_capacity]:
            brand_dict[brand][model][engine_capacity][engine] = hp

    with open('brand_dict.json', 'w', encoding='utf-8') as f:
        json.dump(brand_dict, f, ensure_ascii=False)

    new_df = pd.DataFrame()
    prev_brand = None
    prev_model = None
    for brand in brand_dict.keys():
        for model in brand_dict[brand].keys():
            start_date = brand_dict[brand][model]["start_date"]
            end_date = brand_dict[brand][model]["end_date"]
            for engine_cap in brand_dict[brand][model].keys():

                if isinstance(brand_dict[brand][model][engine_cap], dict):
                    engine_list = [engines for engines in brand_dict[brand][model][engine_cap].keys()]
                    hp = [brand_dict[brand][model][engine_cap][hp] for hp in brand_dict[brand][model][engine_cap].keys()]
                    if brand != prev_brand:
                        brand_value = brand
                    else:
                        brand_value = ''
                    if model != prev_model:
                        start_date = start_date[-2:] + "." + start_date[:-2]
                        end_date = end_date[-2:] + "." + end_date[:-2]
                        model_value = f'{model} {start_date}-{end_date}'
                    else:
                        model_value = ''
                    df = pd.DataFrame.from_dict({'МОДЕЛЬ': [brand_value, model_value, engine_cap],
                                                 'КОД ДВИГАТЕЛЯ': ['', '', ', '.join(engine_list)],
                                                 'Мощность Л.С': ['', '', ', '.join(hp)],
                                                 'Name': ['', '', brand],
                                                 'VM': ['', '', model],
                                                 'TypeName': ['', '', engine_cap],
                                                 }, orient='index')
                    df = df.transpose()
                    new_df = pd.concat([new_df, df], ignore_index=True)
                    prev_brand = brand
                    prev_model = model
    pd.set_option('display.width', 500)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    out_df = pd.read_excel('result_colum_category.xlsx')
    out_df = out_df.drop(columns=["HorsePowers"])
    out_df = out_df.drop_duplicates(subset=['Name', 'VM', 'TypeName'])

    new_df = new_df.merge(out_df, how='left', on=["Name", "VM", "TypeName"])
    new_df = new_df.drop(columns=["Name"])
    new_df = new_df.drop(columns=["VM"])
    new_df = new_df.drop(columns=["TypeName"])
    new_df = new_df.drop(columns=["Engines"])
    new_df.rename(columns={'Фильтр, воздух во внутренном пространстве': 'Салонный фильтр'}, inplace=True)
    new_df = new_df.drop(columns=["Year"])
    new_df = new_df[new_df['МОДЕЛЬ'] != '']
    print(new_df.head(10))
    new_df.to_excel("result_vehicle_sort.xlsx", index=False)


def strip_filter():
    df = pd.read_excel('result_vehicle_sort.xlsx')
    df['Салонный фильтр CAC'] = df['Салонный фильтр'].apply(lambda x: ', '.join([val.strip() for val in str(x).split(',')
                                                                                 if str(val).strip().startswith('CAC')]))
    df['Салонный фильтр CAB'] = df['Салонный фильтр'].apply(lambda x: ', '.join([val.strip() for val in str(x).split(',')
                                                                                 if str(val).strip().startswith('CAB')]))
    df['Салонный фильтр CA'] = df['Салонный фильтр'].apply(lambda x: ', '.join([val.strip() for val in str(x).split(',')
                                                                                if str(val).strip().startswith('CA')
                                                                                and not str(val).strip().startswith('CAC')
                                                                                and not str(val).strip().startswith('CAB')]))
    df = df.drop('Салонный фильтр', axis=1)
    # print(df)
    df.to_excel("res_strip_filter.xlsx", index=False)



def color_rows(input_file):
    wb = load_workbook(filename=input_file)
    ws = wb.active

    for i in range(2, ws.max_row + 1):
        current_row = ws[i]
        prev_row = ws[i - 1]

        if not current_row[1].value and not prev_row[1].value:
            for cell in current_row:
                cell.fill = PatternFill(start_color="FF8300", end_color="FF8300", fill_type="solid")  # оранжевый

            for cell in prev_row:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # красный

        elif not current_row[1].value:
            for cell in current_row:
                cell.fill = PatternFill(start_color="FF8300", end_color="FF8300", fill_type="solid")  # оранжевый
        else:
            for cell in current_row:
                cell.fill = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")  # бежевый

        # print('')
    output_file = input_file.replace('.xlsx', 'result.xlsx')
    wb.save(output_file)



def art_criteria():
    df = pd.read_excel('res_strip_filterresult.xlsx')

    # Создаем копию DataFrame
    df_copy = df.copy()

    # Разделяем значения и создаем новую строку
    df_copy['Радиатор, охлаждение двигателя'] = df_copy['Радиатор, охлаждение двигателя'].str.split(',')
    # Разворачиваем списки в отдельные строки
    df_copy = df_copy.shift(periods=2).explode('Радиатор, охлаждение двигателя', ignore_index=True)

    # Разделяем значения и создаем новую строку
    df_copy['Масляный фильтр'] = df_copy['Масляный фильтр'].str.split(',')
    # Разворачиваем списки в отдельные строки
    df_copy = df_copy.shift(periods=2).explode('Масляный фильтр', ignore_index=True)

    # Разделяем значения и создаем новую строку
    df_copy['Топливный фильтр'] = df_copy['Топливный фильтр'].str.split(',')
    # Разворачиваем списки в отдельные строки
    df_copy = df_copy.shift(periods=2).explode('Топливный фильтр', ignore_index=True)

    # Разделяем значения и создаем новую строку
    df_copy['Воздушный фильтр'] = df_copy['Воздушный фильтр'].str.split(',')
    # Разворачиваем списки в отдельные строки
    df_copy = df_copy.shift(periods=2).explode('Воздушный фильтр', ignore_index=True)

    # Разделяем значения и создаем новую строку
    df_copy['Спортивный воздушный фильтр'] = df_copy['Спортивный воздушный фильтр'].str.split(',')
    # Разворачиваем списки в отдельные строки
    df_copy = df_copy.shift(periods=2).explode('Спортивный воздушный фильтр', ignore_index=True)

    # Разделяем значения и создаем новую строку
    df_copy['Салонный фильтр CAC'] = df_copy['Салонный фильтр CAC'].str.split(',')
    # Разворачиваем списки в отдельные строки
    df_copy = df_copy.shift(periods=2).explode('Салонный фильтр CAC', ignore_index=True)

    # Разделяем значения и создаем новую строку
    df_copy['Салонный фильтр CAB'] = df_copy['Салонный фильтр CAB'].str.split(',')
    # Разворачиваем списки в отдельные строки
    df_copy = df_copy.shift(periods=2).explode('Салонный фильтр CAB', ignore_index=True)

    # Разделяем значения и создаем новую строку
    df_copy['Салонный фильтр CA'] = df_copy['Салонный фильтр CA'].str.split(',')
    # Разворачиваем списки в отдельные строки
    df_copy = df_copy.shift(periods=2).explode('Салонный фильтр CA', ignore_index=True)

    f1 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Масляный фильтр'], keep='first')
    df_copy.loc[f1, 'Масляный фильтр'] = None

    f2 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Топливный фильтр'], keep='first')
    df_copy.loc[f2, 'Топливный фильтр'] = None

    f3 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Воздушный фильтр'], keep='first')
    df_copy.loc[f3, 'Воздушный фильтр'] = None

    f4 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Спортивный воздушный фильтр'], keep='first')
    df_copy.loc[f4, 'Спортивный воздушный фильтр'] = None

    f5 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Радиатор, охлаждение двигателя'], keep='first')
    df_copy.loc[f5, 'Радиатор, охлаждение двигателя'] = None

    f6 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Конденсатор, кондиционер'], keep='first')
    df_copy.loc[f6, 'Конденсатор, кондиционер'] = None

    f7 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Интеркулер'], keep='first')
    df_copy.loc[f7, 'Интеркулер'] = None

    f8 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Гидрофильтр, автоматическая коробка передач'], keep='first')
    df_copy.loc[f8, 'Гидрофильтр, автоматическая коробка передач'] = None

    f9 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Комплект гидрофильтров, автоматическая коробка передач'], keep='first')
    df_copy.loc[f9, 'Комплект гидрофильтров, автоматическая коробка передач'] = None

    f10 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Топливно-водяной сепаратор'], keep='first')
    df_copy.loc[f10, 'Топливно-водяной сепаратор'] = None

    f11 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Фильтр охлаждающей жидкости'], keep='first')
    df_copy.loc[f11, 'Фильтр охлаждающей жидкости'] = None

    f12 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Патрон осушителя воздуха, пневматическая система'], keep='first')
    df_copy.loc[f12, 'Патрон осушителя воздуха, пневматическая система'] = None

    f13 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Салонный фильтр CAC'], keep='first')
    df_copy.loc[f13, 'Салонный фильтр CAC'] = None

    f14 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Салонный фильтр CAB'], keep='first')
    df_copy.loc[f14, 'Салонный фильтр CAB'] = None

    f15 = df_copy.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Салонный фильтр CA'], keep='first')
    df_copy.loc[f15, 'Салонный фильтр CA'] = None

    df_copy = df_copy.drop_duplicates(
        subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Масляный фильтр', 'Топливный фильтр',
                'Воздушный фильтр', 'Спортивный воздушный фильтр', 'Радиатор, охлаждение двигателя',
                'Конденсатор, кондиционер', 'Интеркулер', 'Гидрофильтр, автоматическая коробка передач',
                'Комплект гидрофильтров, автоматическая коробка передач', 'Топливно-водяной сепаратор',
                'Фильтр охлаждающей жидкости', 'Патрон осушителя воздуха, пневматическая система',
                'Салонный фильтр CAC', 'Салонный фильтр CAB', 'Салонный фильтр CA'], keep=False)

    # Сохраняем в новом файле
    df_copy.to_excel("res_art_criteria.xlsx", index=False)
    print(df_copy)


def add_crit():
    df1 = pd.read_excel('res_art_criteria.xlsx')
    df2 = pd.read_excel('Sacura_export2..xlsx')

    # Convert 'Criterias' column to list of dictionaries
    df2['Criterias'] = df2['Criterias'].apply(ast.literal_eval)

    # Define a function to extract 'key' and 'value' from the list of dictionaries
    def get_key(criteria):
        if criteria and isinstance(criteria, list) and len(criteria) > 0:
            return criteria[0].get('key')
        return None

    def get_value(criteria):
        if criteria and isinstance(criteria, list) and len(criteria) > 0:
            return criteria[0].get('value')
        return None
    df2['NEW_CRIT'] = df2['Criterias'].apply(get_key) + ': ' + df2['Criterias'].apply(get_value)
    df2 = df2[['ArticleNumber', 'NEW_CRIT', 'TypeName']]
    df2.rename(columns={'TypeName': 'VM_df2'}, inplace=True)
    # print(df2)
    df2 = df2.drop_duplicates(subset=['ArticleNumber', 'NEW_CRIT', 'VM_df2'], keep=False)
    # print(df2)

    # Создаем копию основного DataFrame
    merged_1 = df1.copy()
    list_df1 = ['Масляный фильтр', 'МОДЕЛЬ']
    list_dff1 = ['ArticleNumber', 'VM_df2']
    merged_1 = pd.merge(merged_1, df2, left_on=list_df1, right_on=list_dff1, how='left')
    merged_1['Масляный фильтр'] = merged_1['Масляный фильтр'] + ' ' + merged_1['NEW_CRIT'].fillna('')
    merged_1 = merged_1.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_2 = merged_1.copy()
    list_df2 = ['Топливный фильтр', 'МОДЕЛЬ']
    list_dff2 = ['ArticleNumber', 'VM_df2']
    merged_2 = pd.merge(merged_2, df2, left_on=list_df2, right_on=list_dff2, how='left')
    merged_2['Топливный фильтр'] = merged_2['Топливный фильтр'] + ' ' + merged_2['NEW_CRIT'].fillna('')
    merged_2 = merged_2.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_3 = merged_2.copy()
    list_df3 = ['Воздушный фильтр', 'МОДЕЛЬ']
    list_dff3 = ['ArticleNumber', 'VM_df2']
    merged_3 = pd.merge(merged_3, df2, left_on=list_df3, right_on=list_dff3, how='left')
    merged_3['Воздушный фильтр'] = merged_3['Воздушный фильтр'] + ' ' + merged_3['NEW_CRIT'].fillna('')
    merged_3 = merged_3.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_4 = merged_3.copy()
    list_df4 = ['Спортивный воздушный фильтр', 'МОДЕЛЬ']
    list_dff4 = ['ArticleNumber', 'VM_df2']
    merged_4 = pd.merge(merged_4, df2, left_on=list_df4, right_on=list_dff4, how='left')
    merged_4['Спортивный воздушный фильтр'] = merged_4['Спортивный воздушный фильтр'] + ' ' + merged_4['NEW_CRIT'].fillna('')
    merged_4 = merged_4.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_5 = merged_4.copy()
    list_df5 = ['Радиатор, охлаждение двигателя', 'МОДЕЛЬ']
    list_dff5 = ['ArticleNumber', 'VM_df2']
    merged_5 = pd.merge(merged_5, df2, left_on=list_df5, right_on=list_dff5, how='left')
    merged_5['Радиатор, охлаждение двигателя'] = merged_5['Радиатор, охлаждение двигателя'] + ' ' + merged_5['NEW_CRIT'].fillna('')
    merged_5 = merged_5.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_6 = merged_5.copy()
    list_df6 = ['Конденсатор, кондиционер', 'МОДЕЛЬ']
    list_dff6 = ['ArticleNumber', 'VM_df2']
    merged_6 = pd.merge(merged_6, df2, left_on=list_df6, right_on=list_dff6, how='left')
    merged_6['Конденсатор, кондиционер'] = merged_6['Конденсатор, кондиционер'] + ' ' + merged_6['NEW_CRIT'].fillna('')
    merged_6 = merged_6.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_7 = merged_6.copy()
    list_df7 = ['Интеркулер', 'МОДЕЛЬ']
    list_dff7 = ['ArticleNumber', 'VM_df2']
    merged_7 = pd.merge(merged_7, df2, left_on=list_df7, right_on=list_dff7, how='left')
    merged_7['Интеркулер'] = merged_7['Интеркулер'] + ' ' + merged_7['NEW_CRIT'].fillna('')
    merged_7 = merged_7.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_8 = merged_7.copy()
    list_df8 = ['Гидрофильтр, автоматическая коробка передач', 'МОДЕЛЬ']
    list_dff8 = ['ArticleNumber', 'VM_df2']
    merged_8 = pd.merge(merged_8, df2, left_on=list_df8, right_on=list_dff8, how='left')
    merged_8['Гидрофильтр, автоматическая коробка передач'] = merged_8['Гидрофильтр, автоматическая коробка передач'] + ' ' + merged_8['NEW_CRIT'].fillna('')
    merged_8 = merged_8.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_9 = merged_8.copy()
    list_df9 = ['Комплект гидрофильтров, автоматическая коробка передач', 'МОДЕЛЬ']
    list_dff9 = ['ArticleNumber', 'VM_df2']
    merged_9 = pd.merge(merged_9, df2, left_on=list_df9, right_on=list_dff9, how='left')
    merged_9['Комплект гидрофильтров, автоматическая коробка передач'] = merged_9['Комплект гидрофильтров, автоматическая коробка передач'] + ' ' + merged_9['NEW_CRIT'].fillna('')
    merged_9 = merged_9.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_10 = merged_9.copy()
    list_df10 = ['Топливно-водяной сепаратор', 'МОДЕЛЬ']
    list_dff10 = ['ArticleNumber', 'VM_df2']
    merged_10 = pd.merge(merged_10, df2, left_on=list_df10, right_on=list_dff10, how='left')
    merged_10['Топливно-водяной сепаратор'] = merged_10['Топливно-водяной сепаратор'] + ' ' + merged_10['NEW_CRIT'].fillna('')
    merged_10 = merged_10.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_11 = merged_10.copy()
    list_df11 = ['Фильтр охлаждающей жидкости', 'МОДЕЛЬ']
    list_dff11 = ['ArticleNumber', 'VM_df2']
    merged_11 = pd.merge(merged_11, df2, left_on=list_df11, right_on=list_dff11, how='left')
    merged_11['Фильтр охлаждающей жидкости'] = merged_11['Фильтр охлаждающей жидкости'] + ' ' + merged_11['NEW_CRIT'].fillna('')
    merged_11 = merged_11.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_12 = merged_11.copy()
    list_df12 = ['Патрон осушителя воздуха, пневматическая система', 'МОДЕЛЬ']
    list_dff12 = ['ArticleNumber', 'VM_df2']
    merged_12 = pd.merge(merged_12, df2, left_on=list_df12, right_on=list_dff12, how='left')
    merged_12['Патрон осушителя воздуха, пневматическая система'] = merged_12['Патрон осушителя воздуха, пневматическая система'] + ' ' + merged_12['NEW_CRIT'].fillna('')
    merged_12 = merged_12.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_13 = merged_12.copy()
    list_df13 = ['Салонный фильтр CAC', 'МОДЕЛЬ']
    list_dff13 = ['ArticleNumber', 'VM_df2']
    merged_13 = pd.merge(merged_13, df2, left_on=list_df13, right_on=list_dff13, how='left')
    merged_13['Салонный фильтр CAC'] = merged_13['Салонный фильтр CAC'] + ' ' + merged_13['NEW_CRIT'].fillna('')
    merged_13 = merged_13.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_14 = merged_13.copy()
    list_df14 = ['Салонный фильтр CAB', 'МОДЕЛЬ']
    list_dff14 = ['ArticleNumber', 'VM_df2']
    merged_14 = pd.merge(merged_14, df2, left_on=list_df14, right_on=list_dff14, how='left')
    merged_14['Салонный фильтр CAB'] = merged_14['Салонный фильтр CAB'] + ' ' + merged_14['NEW_CRIT'].fillna('')
    merged_14 = merged_14.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])

    merged_15 = merged_14.copy()
    list_df15 = ['Салонный фильтр CA', 'МОДЕЛЬ']
    list_dff15 = ['ArticleNumber', 'VM_df2']
    merged_15 = pd.merge(merged_15, df2, left_on=list_df15, right_on=list_dff15, how='left')
    merged_15['Салонный фильтр CA'] = merged_15['Салонный фильтр CA'] + ' ' + merged_15['NEW_CRIT'].fillna('')
    merged_15 = merged_15.drop(columns=['ArticleNumber', 'NEW_CRIT', 'VM_df2'])


    merged_df = merged_15[['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Масляный фильтр', 'Топливный фильтр',
                'Воздушный фильтр', 'Спортивный воздушный фильтр', 'Радиатор, охлаждение двигателя',
                'Конденсатор, кондиционер', 'Интеркулер', 'Гидрофильтр, автоматическая коробка передач',
                'Комплект гидрофильтров, автоматическая коробка передач', 'Топливно-водяной сепаратор',
                'Фильтр охлаждающей жидкости', 'Патрон осушителя воздуха, пневматическая система',
                'Салонный фильтр CAC', 'Салонный фильтр CAB', 'Салонный фильтр CA']]

    f1 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Масляный фильтр'], keep='first')
    merged_df.loc[f1, 'Масляный фильтр'] = None

    f2 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Топливный фильтр'], keep='first')
    merged_df.loc[f2, 'Топливный фильтр'] = None

    f3 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Воздушный фильтр'], keep='first')
    merged_df.loc[f3, 'Воздушный фильтр'] = None

    f4 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Спортивный воздушный фильтр'],
                            keep='first')
    merged_df.loc[f4, 'Спортивный воздушный фильтр'] = None

    f5 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Радиатор, охлаждение двигателя'],
                            keep='first')
    merged_df.loc[f5, 'Радиатор, охлаждение двигателя'] = None

    f6 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Конденсатор, кондиционер'],
                            keep='first')
    merged_df.loc[f6, 'Конденсатор, кондиционер'] = None

    f7 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Интеркулер'], keep='first')
    merged_df.loc[f7, 'Интеркулер'] = None

    f8 = merged_df.duplicated(
        subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Гидрофильтр, автоматическая коробка передач'], keep='first')
    merged_df.loc[f8, 'Гидрофильтр, автоматическая коробка передач'] = None

    f9 = merged_df.duplicated(
        subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Комплект гидрофильтров, автоматическая коробка передач'],
        keep='first')
    merged_df.loc[f9, 'Комплект гидрофильтров, автоматическая коробка передач'] = None

    f10 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Топливно-водяной сепаратор'],
                             keep='first')
    merged_df.loc[f10, 'Топливно-водяной сепаратор'] = None

    f11 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Фильтр охлаждающей жидкости'],
                             keep='first')
    merged_df.loc[f11, 'Фильтр охлаждающей жидкости'] = None

    f12 = merged_df.duplicated(
        subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Патрон осушителя воздуха, пневматическая система'],
        keep='first')
    merged_df.loc[f12, 'Патрон осушителя воздуха, пневматическая система'] = None

    f13 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Салонный фильтр CAC'], keep='first')
    merged_df.loc[f13, 'Салонный фильтр CAC'] = None

    f14 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Салонный фильтр CAB'], keep='first')
    merged_df.loc[f14, 'Салонный фильтр CAB'] = None

    f15 = merged_df.duplicated(subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Салонный фильтр CA'], keep='first')
    merged_df.loc[f15, 'Салонный фильтр CA'] = None

    merged_df = merged_df.drop_duplicates(
        subset=['МОДЕЛЬ', 'КОД ДВИГАТЕЛЯ', 'Мощность Л.С', 'Масляный фильтр', 'Топливный фильтр',
                'Воздушный фильтр', 'Спортивный воздушный фильтр', 'Радиатор, охлаждение двигателя',
                'Конденсатор, кондиционер', 'Интеркулер', 'Гидрофильтр, автоматическая коробка передач',
                'Комплект гидрофильтров, автоматическая коробка передач', 'Топливно-водяной сепаратор',
                'Фильтр охлаждающей жидкости', 'Патрон осушителя воздуха, пневматическая система',
                'Салонный фильтр CAC', 'Салонный фильтр CAB', 'Салонный фильтр CA'], keep='first')

    merged_df.to_excel("test.xlsx", index=False)
    print(merged_df)



if __name__ == '__main__':
    get_category()
    # change_colum()
    # strip_filter()
    #
    # art_criteria()

    # add_crit()

    # color_rows('res_art_criteria.xlsx')
